import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

excel_file = "Output.xlsx"


def save_to_excel(data):
    df = pd.DataFrame(data)

    # Check if Excel file exists before writing
    if os.path.exists(excel_file):
        print(f"Excel file '{excel_file}' exists. Overwriting...")

    # Save DataFrame to Excel
    df.to_excel(excel_file, index=False, sheet_name="Sheet1")

    # Apply formatting if file was created or overwritten
    wb = load_workbook(excel_file)
    ws = wb.active

    # Bold the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Add padding

    # Save the formatted Excel file
    wb.save(excel_file)
    print(f"Formatted JSON data saved to {excel_file} successfully!")
